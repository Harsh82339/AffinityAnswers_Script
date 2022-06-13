import pandas as pd
import openpyxl

#------------Initialization---------------------------------------------------
wb = openpyxl.load_workbook('realdonaldtrump.xlsx')                   # we define the input sheet | Assumption same Folder as script
sheet = wb.active                                                     #define the sheet name
patterns =['facebook','youtube']                                      #Key Words that we are focusing on
Id_Column = 1                                                         # Unique Id Column
Content_Column = 3                                                    # Content Column
#--------------------------------------------------------------------------------

                                                                      #iterating each Row, one by one
for r in range(2,sheet.max_row+1):

    wb_id = sheet.cell(row = r, column = Id_Column).value             #Storing the unique Id for Each Tweet | Record 
    wb_content = sheet.cell(row = r, column = Content_Column).value   #reading the tweets of each person
    
                                                                      #iterating to check our focuse words/key
    for pattern in patterns:
        if pattern in wb_content:
            print("Racial Slurs |", wb_id, "|", pattern)              #Print the Id's those are matching focus keywords, along with the keyword
